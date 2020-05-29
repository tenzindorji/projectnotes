import openpyxl
import xlrd

theFile=xlrd.open_workbook('/pythongLearning/awscost.xls')

#theFile = openpyxl.load_workbook('/pythongLearning/awscost.xls')
allSheetNames = theFile.sheetnames

print("All sheet names {} " .format(theFile.sheetnames))


def find_specific_cell():
    for row in range(1, currentSheet.max_row + 1):
        for column in "Billing Component":  # Here you can add or reduce the columns
            cell_name = "{}{}".format(column, row)
            if currentSheet[cell_name].value == "tdi":
                #print("{1} cell is located on {0}" .format(cell_name, currentSheet[cell_name].value))
                print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))
                return cell_name

def get_column_letter(specificCellLetter):
    letter = specificCellLetter[0:-1]
    print(letter)
    return letter

def get_all_values_by_cell_letter(letter):
    for row in range(1, currentSheet.max_row + 1):
        for column in letter:
            cell_name = "{}{}".format(column, row)
            #print(cell_name)
            print("cell position {} has value {}".format(cell_name, currentSheet[cell_name].value))



for sheet in allSheetNames:
    print("Current sheet name is {}" .format(sheet))
    currentSheet = theFile[sheet]
    specificCellLetter = (find_specific_cell())
    letter = get_column_letter(specificCellLetter)

    get_all_values_by_cell_letter(letter)

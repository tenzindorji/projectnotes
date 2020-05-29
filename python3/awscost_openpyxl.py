# Reading an excel file using Python
import xlrd
import openpyxl
from openpyxl.workbook import Workbook
import sys
import yaml
import re
import pyexcel as p

# Give the location of the file

#Depedency:
'''
pip3 install pyexcel-xls
pip3 install pyexcel-xlsx
pip3 install pyexcel-xlsxw
'''

def extract_data(filename):


    p.save_book_as(file_name=filename, dest_file_name='awscost_converted.xlsx')

    #load excel input file
    '''
    wb=xlrd.open_workbook(filename)
    sheet = wb.sheet_by_index(0)
    sheet.cell_value(0, 0)
    '''
    wb=openpyxl.load_workbook('/pythongLearning/awscost_converted.xlsx')
    workbook='/pythongLearning/awscost_converted.xlsx'
    ws=wb.active
    #first_row=ws[1]
    print(ws.max_row)
    print(ws.max_column)


    #load work book


    #headers       = ['Company','Address','Tel','Web']
    headers= ["Account #", "Acct. Name", "Period", "Financial BUFG", "BUFG L2 (Ops)", "BUFG L3 (Ops)", "Application", "Usage", "App Env", "Billing BU", "Billing Component", "Billing fp", "Billing User", "Billing User App", "Cost", "Directs"]
    #create new xlsx
    workbook_name = 'awscostdata.xlsx'
    wwb = Workbook()
    page = wwb.active
    page.title = 'awscostdata'
    page.append(headers)

    #directs={}
    #load service owners Bhushan's Direct
    directs=yaml.load(open('/pythongLearning/directs.txt'))
    #print(y)

    r=ws.max_row
    for index in range(r):
        print(index)
        for row in ws.values:
            #if "acc" in row:
            #    print(row)
            ws.delete_rows(index,1)
                #print(service)
            r=ws.max_row
    wb.save(filename='/pythongLearning/awscost_converted.xlsx')






    wwb.save(filename=workbook_name)
    #print(sheet.ncols)


def main():

    #if not input_file:
    if len(sys.argv) != 2:
        print ('Provide excel file as input\nUsage: python3', sys.argv[0], 'excel_filename')
        sys.exit(1)

    input_file=sys.argv[1]
    finalexceloutput=extract_data(input_file)
    print (finalexceloutput)

if __name__=='__main__':
    main()
